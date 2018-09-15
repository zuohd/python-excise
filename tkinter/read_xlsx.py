from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    wb=load_workbook(filename=path)
    # print(wb.sheetnames)
    sheet=wb[wb.sheetnames[0]]
    # print(sheet.max_row)
    # print(sheet.max_column)
    # print(sheet.title)
    for lineNum in range(1,sheet.max_row+1):
        # print(lineNum)
        linList=[]
        for columnNum in range(1,sheet.max_column+1):
            value=sheet.cell(row=lineNum,column=columnNum).value
            if value!=None:
                linList.append(value)
            else:
                break
        print(linList)

readXlsxFile("test.xlsx")